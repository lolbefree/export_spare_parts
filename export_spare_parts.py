#!/usr/bin/python
# -*- coding: utf-8 -*-
from PyQt5 import QtWidgets, uic
import pyodbc
import openpyxl
from openpyxl import Workbook
import sys
from untitled import Ui_Export_spare_parts
import sql_querys


class SpareParts(QtWidgets.QDialog):
    wb = Workbook()
    server = ''
    database = ''
    username = ''
    password = 'PW'
    driver = '{SQL Server}'  # Driver you need to connect to the database
    port = '1433'

    def __init__(self):
        self.ui = Ui_Export_spare_parts()
        super().__init__()
        self.ui.setupUi(self)
        self.provider_link_before_delete = ""
        self.ui.open_excel_button.clicked.connect(lambda x: self.showDialog())
        self.ui.commandLinkButton.clicked.connect(lambda x: self.check_in_base())
        self.ui.pushButton.clicked.connect(lambda x: self.foresight())
        self.ui.del_from_iprr.clicked.connect(lambda x: self.delete_from_iprr_and_iprh())
        self.show()  # Show the GUI
        self.cnn = pyodbc.connect(
            'DRIVER=' + self.driver + ';PORT=port;SERVER=' + self.server + ';PORT=1443;DATABASE=' + self.database + ';UID=' + self.username +
            ';PWD=' + self.password)

        self.group_err = False
        self.discount_err = False
        self.code_list = list()
        self.name_list = list()
        self.group_list = list()
        self.discount_list = list()
        self.enter_price_list = list()
        self.retail_list = list()
        self.provider_list = list()
        self.original_code = list()  # Список оригинальных кодов с спец символами
        self.original_code_without_symbols = list()  # Список оригинальных кодов без пец символами
        self.sql_iprr = str()
        self.sql_iprh = str()
        self.main_dict = dict()
        self.cursor = self.cnn.cursor()
        self.cnt = 0
        self.if_exist_in_base = str()
        self.ui.pushButton_2.clicked.connect(lambda x: self.add_to_main_base())

    def delete_from_iprr_and_iprh(self):
        self.cursor.execute(sql_querys.delelet_from_iprr(self.provider_link_before_delete))
        self.cnn.commit()
        self.cursor.execute(sql_querys.delelet_from_iprh(self.provider_link_before_delete))
        self.cnn.commit()
        self.ui.print_res.setText("Данный каталог удален с IPRR")
        self.ui.print_res.setStyleSheet("color: blue")

    def clear_lists_of_data(self):
        self.ui.print_res.setText("")
        for data in [self.original_code_without_symbols, self.name_list, self.group_list, self.discount_list,
                     self.enter_price_list,
                     self.retail_list, self.provider_list, self.original_code]:
            data.clear()

    def check_float(self, potential_float):
        try:
            float(potential_float)
            return True
        except ValueError:
            return False

    def create_list(self, later, row_num, name_of_list):
        # print(later, row_num, name_of_list)
        self.wb = openpyxl.load_workbook(self.filename, data_only=True)
        self.ws = self.wb[self.ui.later.text()]

        row_max = self.ws.max_row  # не забыть отнять 1

        if later == row_num:
            if name_of_list == "group_" and self.check_float(self.ui.group_.text()):
                self.group_list.append(self.ui.group_.text())
                self.group_list = self.group_list * (row_max - 1)
                self.group_err = False
            if name_of_list == "group_" and not self.check_float(self.ui.group_.text()):
                self.group_err = True
                self.group_list.append(self.ui.group_.text())
                self.group_list = self.group_list * (row_max - 1)

            if name_of_list == "discount_" and self.check_float(self.ui.discount_.text()):
                self.discount_list.append(self.ui.discount_.text())
                self.discount_list = self.discount_list * (row_max - 1)
                self.discount_err = False
            if name_of_list == "discount_" and not self.check_float(self.ui.discount_.text()):
                self.discount_err = True
                self.discount_list.append(self.ui.discount_.text())
                self.discount_list = self.discount_list * (row_max - 1)
            if name_of_list == "provider_":
                self.provider_list.append(self.ui.provider_.text())
                self.provider_list = self.provider_list * (row_max - 1)

        else:
            if not self.ui.disccount_check.isChecked():
                self.discount_err = False
            if not self.ui.group_check.isChecked():
                self.group_err = False

            while int(row_num) <= row_max and self.ws[f"{later}{row_num}"].value is not None:
                # print(self.ws[f"{later}{row_num}"].value)
                string = ""
                if name_of_list == "code_":
                    for item in str(self.ws[f"{later}{row_num}"].value):
                        if item.isalpha() or item.isdigit():
                            string += item
                    if int(row_num) <= row_max:
                        self.original_code.append(self.ws[f"{later}{row_num}"].value)
                        self.original_code_without_symbols.append(string)
                elif name_of_list == "name_":
                    self.name_list.append(self.ws[f"{later}{row_num}"].value)
                elif name_of_list == "group_" and not self.ui.group_check.isChecked():
                    self.group_list.append(self.ws[f"{later}{row_num}"].value)
                elif name_of_list == "enter_price":
                    self.enter_price_list.append(self.ws[f"{later}{row_num}"].value)
                elif name_of_list == "retail_":
                    self.retail_list.append(self.ws[f"{later}{row_num}"].value)
                elif name_of_list == "discount_" and not self.ui.disccount_check.isChecked():
                    self.discount_list.append(self.ws[f"{later}{row_num}"].value)
                elif name_of_list == "provider_" and not self.ui.provider_check.isChecked():
                    self.provider_list.append(self.ws[f"{later}{row_num}"].value)
                row_num = int(row_num) + 1

    def showDialog(self):
        fname = QtWidgets.QFileDialog.getOpenFileName(self, 'Open file', '*.xlsx')[0]
        name_index_ = fname.rfind("/")
        self.filename = fname
        self.ui.label_7.setText(fname[name_index_ + 1:])
        self.main_dict.clear()

    def start_main_work(self):
        self.create_list(self.ui.code_.text()[:1], self.ui.code_.text()[1:], "code_")
        self.create_list(self.ui.name_.text()[:1], self.ui.name_.text()[1:], "name_")
        if len(self.ui.group_.text()) > 1 and not self.ui.group_check.isChecked():
            self.create_list(self.ui.group_.text()[:1], self.ui.group_.text()[1:], "group_")
        else:
            self.create_list(self.ui.group_.text(), self.ui.group_.text(), "group_")

        if len(self.ui.discount_.text()) > 1 and not self.ui.disccount_check.isChecked():
            self.create_list(self.ui.discount_.text()[:1], self.ui.discount_.text()[1:], "discount_")
        else:
            self.create_list(self.ui.discount_.text(), self.ui.discount_.text(), "discount_")
        self.create_list(self.ui.enter_price.text()[:1], self.ui.enter_price.text()[1:], "enter_price")
        self.create_list(self.ui.retail_.text()[:1], self.ui.retail_.text()[1:], "retail_")
        if len(self.ui.provider_.text()) > 1 and not self.ui.provider_check.isChecked():

            self.create_list(self.ui.provider_.text()[:1], self.ui.provider_.text()[1:], "provider_")
        else:
            self.create_list(self.ui.provider_.text(), self.ui.provider_.text(), "provider_")

    def closeEvent(self, event):
        self.cnn.close()

    def check_in_base(self):
        if_exist_ = f"""
                if exists 
        (select top 1 * from iprr where SUPLNO='{self.provider_link_before_delete}')
        select 'true'
        else 
        select 'false'


                """
        res = self.cursor.execute(if_exist_)
        for row in res:
            if (row[0]) == "true":
                self.if_exist_in_base = "true"
            else:
                self.if_exist_in_base = "false"

        if self.if_exist_in_base == "false":
            self.insert_in_database()
        else:
            self.ui.print_res.setText("Данный каталог уже есть в IPRR")
            self.ui.print_res.setStyleSheet("color: red")

    def insert_in_database(self):
        self.ui.progressBar.setMaximum(len(self.main_dict))

        # try:
        res = self.cursor.execute(sql_querys.check_supl(self.provider_link_before_delete))
        res = list(res)
        for row in res:
            if (row[0]) == "true":
                self.cursor.execute(sql_querys.suplno_config(self.provider_link_before_delete))
                self.cnn.commit()
                self.sql_iprh = f"""
                INSERT INTO iprh (created,CTYPE,PRSETDT,SUPLNO,USRSID,UPDPAC,UPDSAL,CRENEW,CHGLIS,NEWLIS,ONLYBPR,CURRCD,CHALIS,CHELIS,NOTE,BPRLIS,FLANG1)
                         values (getdate(),'f',convert(date,getdate()),'{self.provider_link_before_delete}','auto',1,0,1,0,0,0,'uah',0,0,'{"OK " + str(len(self.main_dict))}',0,'eng')"""
                # print(self.sql_iprh)
                self.cursor.execute(self.sql_iprh)
                self.sql_iprr_key = f"""declare @key datetime

                set @key=(select max(created) from iprh where SUPLNO='{self.provider_link_before_delete}'
                         group by SUPLNO)
                select @key"""

                for row in self.cursor.execute(self.sql_iprr_key):
                    self.key = row[0]

                self.cnn.commit()
                print("tyt1")
                for ITEMNO in self.main_dict:

                    if "'" in self.main_dict[ITEMNO]["name"]:
                        self.main_dict[ITEMNO]["name"] = self.main_dict[ITEMNO]["name"].replace("'", "`")
                    group_id = f"""select CONVERT(varchar,convert(integer,IGROUPID), 100)  from igrp
where SUPLNO='{self.main_dict[ITEMNO]["SUPLNO"]}' and igrpid='{self.main_dict[ITEMNO]["IGRPID"]}'"""
                    print(f"group_id: {group_id}")

                    print(list(self.cursor.execute(group_id)))

                    self.sql_iprr = f"""
                    insert into iprr (CREATED,SUPLNO,ITEMNO,skey,name,SWENAME,IGRPID,DDISCCD,svatcd,BUYPR,SELPR,CURRCD)
                    values (convert(datetime, '{self.key.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]}'),'{self.main_dict[ITEMNO]["SUPLNO"]}','{self.main_dict[ITEMNO]["ITEMNO"]}','{self.main_dict[ITEMNO]["skey"]}','{self.main_dict[ITEMNO]["name"]}','{self.main_dict[ITEMNO]["name"]}','{list(self.cursor.execute(group_id))[0][0]}','{self.main_dict[ITEMNO]["DDISCCD"]}','1' ,{self.main_dict[ITEMNO]["BUYPR"]},{self.main_dict[ITEMNO]["SELPR"]},'UAH')"""
)
                    self.cursor.execute(self.sql_iprr)
                    self.cnn.commit()
                    self.ui.label_itemno.setText(f"Код запчасти : {self.main_dict[ITEMNO]['ITEMNO']}")
                    self.cnt += 1
                    self.ui.progressBar.setValue(self.cnt)

                    self.ui.print_res.setText("Экспорт удачно выполнен!")
                    self.ui.print_res.setStyleSheet("color: green")
                self.main_dict.clear()

            else:

                self.ui.print_res.setText("Сначала добавьте поставщика!")
                self.ui.print_res.setStyleSheet("color: red")


    def foresight_clear(self):

        for data in [self.code_list_foresigh, self.name_list_foresigh, self.group_list_foresigh,
                     self.discount_list_foresigh, self.enter_price_list_foresigh, self.retail_list_foresigh,
                     self.provider_list_foresigh]:
            for item in data:
                item.setText("")

    def main_procedure(self):
        self.start_main_work()

    def clear_all_lists(self):
        try:
            self.provider_link_before_delete = self.provider_list[0]
        except IndexError:
            self.ui.print_res.setText("Проверьте правильность координат")
            self.ui.print_res.setStyleSheet('color: red')
        self.original_code_without_symbols.clear()
        self.original_code.clear()
        self.provider_list.clear()
        self.group_list.clear()
        self.discount_list.clear()
        self.retail_list.clear()
        self.enter_price_list.clear()
        self.name_list.clear()

    def foresight(self):

        try:
            self.code_list_foresigh = [self.ui.code, self.ui.code_2, self.ui.code_3, self.ui.code_4]
            self.name_list_foresigh = [self.ui.name, self.ui.name_2, self.ui.name_3, self.ui.name_4]
            self.group_list_foresigh = [self.ui.group, self.ui.group_2, self.ui.group_3, self.ui.group_4]
            self.discount_list_foresigh = [self.ui.discount, self.ui.discount_2, self.ui.discount_3,
                                           self.ui.discount_4]
            self.enter_price_list_foresigh = [self.ui.enter_price_l, self.ui.enter_price_l_2, self.ui.enter_price_l_3,
                                              self.ui.enter_price_l_4]
            self.retail_list_foresigh = [self.ui.retail, self.ui.retail_2, self.ui.retail_3, self.ui.retail_4]
            self.provider_list_foresigh = [self.ui.provider_1, self.ui.provider_2, self.ui.provider_3,
                                           self.ui.provider_4, ]
            self.foresight_clear()

            self.main_procedure()
            if len(self.original_code_without_symbols) == len(self.name_list) == len(self.group_list) == len(
                    self.discount_list) == len(
                self.enter_price_list) == len(self.retail_list) == len(self.provider_list):
                if len(set(self.provider_list)) > 1:
                    self.ui.print_res.setText("Проверьте колонку поставщика, далжно иметь уникальное значение.")
                    self.ui.print_res.setStyleSheet("color: red")
                else:
                    if len(self.original_code_without_symbols) < 4:
                        len_foresight_print = len(self.original_code_without_symbols)
                    else:
                        len_foresight_print = 4
                    for item in range(len_foresight_print):
                        self.code_list_foresigh[item].setText(str(self.original_code_without_symbols[item]))
                        self.name_list_foresigh[item].setText(str(self.name_list[item]))
                        self.group_list_foresigh[item].setText(str(self.group_list[item]))
                        self.discount_list_foresigh[item].setText(str(self.discount_list[item]))
                        self.enter_price_list_foresigh[item].setText(str(self.enter_price_list[item]))
                        self.retail_list_foresigh[item].setText(str(self.retail_list[item]))
                        self.provider_list_foresigh[item].setText(str(self.provider_list[item]))

                    self.ui.print_res.setStyleSheet('color: green')
                    if self.group_err:
                        self.ui.print_res.setText("Группа должна быть целым числом")
                        self.ui.print_res.setStyleSheet('color: red')
                    elif self.discount_err:
                        self.ui.print_res.setText("Скидка должна быть вещественным или целым числом")
                        self.ui.print_res.setStyleSheet("color: red")
                    elif self.discount_err and self.group_err:
                        self.ui.print_res.setText("Скидка и группа должна быть вещественным или целым числом")
                        self.ui.print_res.setStyleSheet("color: red")
                    else:
                        self.ui.print_res.setText("Предосмотр сформирован")
                    print(self.original_code_without_symbols)
                    for ITEMNO in self.original_code_without_symbols:
                        d = {ITEMNO: {"ITEMNO": 0, "SUPLNO": 0, "name": 0, "IGRPID": 0, "BUYPR": 0, "SELPR": 0,
                                      "skey": 0}}
                        self.main_dict.update(d)
                    for ITEMNO, SUPLNO, name, IGRPID, DDISCCD, BUYPR, SELPR, orig in zip(
                            self.original_code_without_symbols,
                            self.provider_list,
                            self.name_list,
                            self.group_list,
                            self.discount_list,
                            self.enter_price_list,
                            self.retail_list,
                            self.original_code):
                        if float(self.main_dict[ITEMNO]["SELPR"]) < float(SELPR):

                            self.main_dict[ITEMNO]["ITEMNO"] = ITEMNO
                            self.main_dict[ITEMNO]["SUPLNO"] = SUPLNO
                            self.main_dict[ITEMNO]["name"] = name
                            self.main_dict[ITEMNO]["IGRPID"] = IGRPID
                            self.main_dict[ITEMNO]["DDISCCD"] = DDISCCD
                            self.main_dict[ITEMNO]["BUYPR"] = BUYPR
                            self.main_dict[ITEMNO]["SELPR"] = SELPR
                            self.main_dict[ITEMNO]["skey"] = orig
                            if self.main_dict[ITEMNO]["ITEMNO"] in self.original_code_without_symbols:
                                ITEMNO_Index = self.original_code_without_symbols.index(
                                    self.main_dict[ITEMNO]["ITEMNO"])
                                self.main_dict[ITEMNO]["skey"] = self.original_code[ITEMNO_Index]
                            elif self.main_dict[ITEMNO]["ITEMNO"] not in self.original_code_without_symbols:
                                self.main_dict[ITEMNO]["skey"] = ITEMNO

                self.clear_all_lists()

        except KeyError:
            self.ui.print_res.setText("Проверте название листа")
            self.ui.print_res.setStyleSheet("color: red")
        except (AttributeError, openpyxl.utils.exceptions.InvalidFileException):
            self.ui.print_res.setText("Выберите excel")
            self.ui.print_res.setStyleSheet('color: red')

    def add_to_main_base(self):
        # print("add to main base")
        try:
            if self.ui.radioButton_current.isChecked():
                self.cursor.execute(sql_querys.main_query(self.provider_link_before_delete))
                self.cnn.commit()
                self.ui.print_res.setText(f"Обновление каталога {self.provider_link_before_delete} успешен")
                self.ui.print_res.setStyleSheet('color: Green')
            elif self.ui.radioButton_vw:

                self.cursor.execute("dbo.vwpriceimport")
                self.ui.print_res.setText(f"Обновление каталога VW успешен")
                self.ui.print_res.setStyleSheet('color: Green')
            elif self.ui.radioButton_dil:
                self.cursor.execute("dbo.dilpriceimport")
                self.ui.print_res.setText(f"Обновление каталога Dil успешен")
                self.ui.print_res.setStyleSheet('color: Green')

            elif self.ui.radioButton_sk:
                self.cursor.execute("dbo.skpriceimport")
                self.ui.print_res.setText(f"Обновление каталога sk успешен")
                self.ui.print_res.setStyleSheet('color: Green')

            elif self.ui.radioButton_sk:
                self.cursor.execute("dbo.nzppriceimport")
                self.ui.print_res.setText(f"Обновление каталога NZP успешен")
                self.ui.print_res.setStyleSheet('color: Green')
        except Exception as err:
            self.ui.print_res.setText(f"{err}")
            self.ui.print_res.setStyleSheet('color: red')


def main():
    app = QtWidgets.QApplication(sys.argv)
    SpareParts().exec_()


if __name__ == '__main__':
    main()
